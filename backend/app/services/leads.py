"""Lead import: spreadsheet parsing, column mapping, validation + dedupe."""
import csv
import io
from typing import Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ..models.campaign_contact import CampaignContact
from ..models.contact import Contact
from ..models import BlacklistedContact


# Canonical column keys a user may map to.
CANONICAL_COLUMNS = ["name", "phone", "company", "email", "website", "city", "country", "notes"]

# Column header aliases (lower-cased, no spaces/punctuation) -> canonical key.
COLUMN_ALIASES = {
    "name": "name",
    "fullname": "name",
    "full_name": "name",
    "firstname": "name",
    "first_name": "name",
    "surname": "name",
    "lastname": "name",
    "last_name": "name",
    "phone": "phone",
    "mobile": "phone",
    "mobilenumber": "phone",
    "number": "phone",
    "whatsapp": "phone",
    "company": "company",
    "organisation": "company",
    "organization": "company",
    "employer": "company",
    "email": "email",
    "e-mail": "email",
    "mail": "email",
    "website": "website",
    "url": "website",
    "web": "website",
    "city": "city",
    "town": "city",
    "country": "country",
    "nation": "country",
    "notes": "notes",
    "remarks": "notes",
    "comments": "notes",
}


def _normalize_key(raw: str) -> str:
    return "".join(ch for ch in (raw or "").strip().lower() if ch.isalnum())


def detect_columns(headers: list[str]) -> dict[str, Optional[str]]:
    """Map each spreadsheet column index to a canonical key (or None)."""
    mapping: dict[str, Optional[str]] = {}
    used: set[str] = set()
    for idx, header in enumerate(headers):
        key = _normalize_key(header)
        canonical = COLUMN_ALIASES.get(key)
        if not canonical:
            # try matching the whole normalized header (handles "phone number" -> "phone")
            canonical = COLUMN_ALIASES.get(_normalize_key(header.replace(" ", "")))
        if canonical and canonical not in used:
            mapping[str(idx)] = canonical
            used.add(canonical)
        else:
            mapping[str(idx)] = None
    return mapping


def _read_rows(stream: io.IOBase, filename: str):
    """Dispatch to the right reader based on extension. Yields (row_number, values)."""
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xls")):
        yield from _read_excel(stream, lower.endswith(".xls"))
    else:
        yield from _read_csv(stream)


def _read_csv(stream):
    data = stream.read() if hasattr(stream, "read") else str(stream).encode()
    text = data.decode("utf-8-sig", "replace") if isinstance(data, (bytes, bytearray)) else str(data)
    sample = text[:65536]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:  # noqa: BLE001
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    for i, row in enumerate(reader):
        if any(cell.strip() for cell in row):
            yield i + 1, row


def _read_excel(stream, xls: bool):
    if xls:
        # xlrd reads the legacy .xls binary format.
        import xlrd  # type: ignore

        book = xlrd.open_workbook(file_contents=stream.read())
        sheet = book.sheet_by_index(0)
        for r in range(sheet.nrows):
            row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
            if any(str(c).strip() for c in row):
                yield r + 1, row
    else:
        from openpyxl import load_workbook  # type: ignore

        stream.seek(0)
        wb = load_workbook(stream, read_only=True, data_only=True)
        sheet = wb.worksheets[0]
        for r, row in enumerate(sheet.iter_rows(values_only=True)):
            if any(str(c).strip() if c is not None else False for c in (row or [])):
                yield r + 1, list(row or [])


# ---------------------------------------------------------------- phone utils
def normalize_phone(raw: str) -> str:
    """Normalize to digits only — matches the convention used by the Contacts CRUD."""
    return "".join(ch for ch in (raw or "") if ch.isdigit())


def is_valid_phone(raw: str) -> bool:
    """Accept E.164-ish or digits-only national numbers (8-15 digits)."""
    norm = normalize_phone(raw)
    return norm.isdigit() and 8 <= len(norm) <= 15


# ---------------------------------------------------------------- parse
def parse_leads(
    file_bytes: bytes, filename: str, column_map: Optional[dict] = None
) -> dict:
    """Parse an uploaded spreadsheet into canonical rows + mapping + stats.

    Returns:
      {
        "columns": [header, ...],          # original headers
        "mapping": {"0": "name", ...},    # detected or user-supplied index->canonical
        "rows":   [{"name":..., "phone":..., "company":..., "extra": {...}}, ...],
        "meta":   {"total": n, "valid": n, "invalid": n, "duplicate": n},
        "headers": [header,...]
      }
    """
    stream = io.BytesIO(file_bytes)
    rows: list[dict] = []
    headers: list[str] = []
    mapping: dict[str, Optional[str]] = {}
    total = 0
    seen_phones: set[str] = set()
    duplicate = 0
    invalid = 0

    reader = _read_rows(stream, filename)
    try:
        first = next(reader)
    except StopIteration:
        return {
            "headers": headers,
            "mapping": mapping,
            "rows": rows,
            "meta": {"total": 0, "valid": 0, "invalid": 0, "duplicate": 0},
        }

    headers = [str(c).strip() if c is not None else "" for c in first[1]]
    mapping = detect_columns(headers)
    if column_map:
        for k, v in column_map.items():
            mapping[str(k)] = v if v in CANONICAL_COLUMNS else None

    for _row_num, raw in reader:
        if total >= 50000:  # hard safety cap per parse
            break
        total += 1
        record: dict = {col: None for col in CANONICAL_COLUMNS}
        record["extra"] = {}
        for idx_str, canonical in mapping.items():
            if not canonical:
                continue
            idx = int(idx_str)
            if idx >= len(raw):
                continue
            val = raw[idx]
            if val is None:
                continue
            val = str(val).strip()
            if not val:
                continue
            if canonical in ("name", "phone", "company", "email", "website", "city", "country", "notes"):
                record[canonical] = val
            else:
                record["extra"][canonical] = val
        phone = normalize_phone(record["phone"] or "")
        if not phone or not is_valid_phone(phone):
            invalid += 1
            record["phone"] = record["phone"] or ""
            record["status"] = "invalid"
            record["reason"] = "Invalid phone number"
            rows.append(record)
            continue
        record["phone"] = phone
        record.pop("extra", None)  # keep extra only if present; harmless
        if phone in seen_phones:
            duplicate += 1
            record["status"] = "duplicate"
            record["reason"] = "Duplicate phone number"
            rows.append(record)
            continue
        record["status"] = "valid"
        record["reason"] = None
        seen_phones.add(phone)
        rows.append(record)

    valid = total - invalid - duplicate
    return {
        "headers": headers,
        "mapping": mapping,
        "rows": rows,
        "meta": {"total": total, "valid": valid, "invalid": invalid, "duplicate": duplicate},
    }


# ---------------------------------------------------------------- rendering
import re as _re

_VAR_RE = _re.compile(r"\{\{\s*([a-zA-Z_][a-zA-Z0-9_]*)(?:\|([^}]+))?\s*\}\}")


def render_template(template: str, contact: dict) -> str:
    """Render {{var|fallback}} placeholders against a contact dict."""
    def repl(m: "_re.Match") -> str:
        key = m.group(1)
        fallback = m.group(2)
        val = contact.get(key) or contact.get("extra", {}).get(key)
        if val not in (None, ""):
            return str(val)
        return (fallback or "").strip()

    return _VAR_RE.sub(repl, template)


# ---------------------------------------------------------------- blacklist / contact lookup
async def is_blacklisted(db: AsyncSession, user_id: str, phone: str) -> bool:
    norm = normalize_phone(phone)
    if not norm:
        return False
    res = await db.scalar(
        select(BlacklistedContact)
        .where(BlacklistedContact.user_id == user_id, BlacklistedContact.phone == norm)
    )
    return res is not None


async def find_existing_contact(db: AsyncSession, user_id: str, phone: str) -> Optional[Contact]:
    """Match an imported lead to an existing CRM contact by normalized phone."""
    norm = normalize_phone(phone)
    if not norm:
        return None
    cand = await db.scalar(
        select(Contact)
        .where(Contact.user_id == user_id, Contact.phone == norm)
    )
    if cand:
        return cand
    # try the digits-only form (contacts may be stored without +)
    digits = norm.lstrip("+")
    return await db.scalar(
        select(Contact)
        .where(Contact.user_id == user_id, Contact.phone == digits)
    )
